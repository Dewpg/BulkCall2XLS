import pandas as pd
import os
import zipfile
import numpy as np
from openpyxl import load_workbook

def handle_irregular_data(filename):
    data = []
    with open(filename, 'r') as f:
        lines = f.readlines()
        headers = [header.replace('"', '') for header in lines[0].strip().split('\t')]
        for line in lines[1:]:
            values = [value.replace('"', '') for value in line.strip().split('\t')]
            row = {headers[i]: values[i] for i in range(len(values))}
            data.append(row)
    return data, headers

# Get the list of zip files in the current directory with "FFIEC CDR" in their name
zip_files = [file for file in os.listdir('.') if "FFIEC CDR" in file and file.endswith('.zip')]

# If there are multiple zip files, query the user for which they would like to use
if len(zip_files) > 1:
    print("Multiple zip files found:")
    for i, file in enumerate(zip_files):
        print(f"{i+1}. {file}")
    file_num = int(input("Enter the number of the file you would like to use: "))
    zip_file = zip_files[file_num - 1]
elif zip_files:
    zip_file = zip_files[0]
else:
    raise Exception("No zip files found with 'FFIEC CDR' in their name.")

# Unzip the file
with zipfile.ZipFile(zip_file, 'r') as zip_ref:
    zip_ref.extractall('extracted')

# Get the list of unzipped files
unzipped_dir = 'extracted'
unzipped_files = os.listdir(unzipped_dir)

# Identify the "Bulk" file and filter out those that have "AL" in their 9th column
bulk_file_name = [file for file in unzipped_files if "Bulk" in file][0]
bulk_file_path = os.path.join(unzipped_dir, bulk_file_name)

# Extract the date from the "Bulk" file name for the sheet name
sheet_name = bulk_file_name.split(' ')[-1].replace('.txt', '')

# Check if the sheet already exists in the Excel file
try:
    book = load_workbook('master.xlsx')
    writer = pd.ExcelWriter('master.xlsx', engine='openpyxl') 
    writer.book = book
    if sheet_name in book.sheetnames:
        replace = input(f"The sheet '{sheet_name}' already exists. Would you like to replace it? (yes/no): ")
        if replace.lower() != 'yes':
            sheet_name += '2'
except FileNotFoundError:
    writer = pd.ExcelWriter('master.xlsx', engine='openpyxl')

# Handle the irregular data in the "Bulk" file
bulk_data, bulk_headers = handle_irregular_data(bulk_file_path)

# Get the list of unique state abbreviations from column 9 of the "Bulk" file
states = list(set(row[bulk_headers[8]] for row in bulk_data))

# Query the user for what State they would like data on
print("States:")
for state in states:
    print(state)
desired_state = input("Enter the state you would like data on: ")

# Store the IDRSSD values from the "Bulk" file in a set for faster lookup
idrssd_set = set(row['IDRSSD'] for row in bulk_data if row[bulk_headers[8]] == desired_state)

# Dictionary to store the headers and corresponding data from each file
data_dict = {}

# Handle the irregular data in the "Bulk" file and store the headers and corresponding data
for row in bulk_data:
    idrssd = row[bulk_headers[0]]
    if idrssd in idrssd_set:
        if idrssd not in data_dict:
            data_dict[idrssd] = {}
        for i in range(1, len(bulk_headers)):
            if bulk_headers[i] in row:
                try:
                    data_dict[idrssd][bulk_headers[i]] = float(row[bulk_headers[i]])
                except ValueError:
                    data_dict[idrssd][bulk_headers[i]] = row[bulk_headers[i]]

# For each file that isn't the README file, the 'Bulk' file, or the "CI" file, store the headers and corresponding data
other_files = [file for file in unzipped_files if "README" not in file and "Bulk" not in file and "CI" not in file]

for file in other_files:
    file_path = os.path.join(unzipped_dir, file)
    
    with open(file_path, 'r') as f:
        headers = [header.replace('"', '') for header in f.readline().strip().split('\t')]

    data = handle_irregular_data(file_path)[0]

    for row in data:
        idrssd = row[headers[0]]
        if idrssd in idrssd_set:
            if idrssd not in data_dict:
                data_dict[idrssd] = {}
            for i in range(1, len(headers)):
                if headers[i] in row:
                    try:
                        data_dict[idrssd][headers[i]] = float(row[headers[i]])
                    except ValueError:
                        data_dict[idrssd][headers[i]] = row[headers[i]]

# Convert the data dictionary into a pandas DataFrame
df = pd.DataFrame(data_dict).T

# Replace empty strings with NaN
df.replace('', np.nan, inplace=True)

# Drop the columns that contain no data
df = df.dropna(axis=1, how='all')

# Write the DataFrame to the Excel file with the specified sheet name
df.to_excel(writer, sheet_name=sheet_name, index_label='IDRSSD')

# Save and close the Excel file
writer.close()
